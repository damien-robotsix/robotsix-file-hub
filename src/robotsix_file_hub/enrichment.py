"""LLM enrichment pipeline: text extraction + llmio chat + embeddings.

Extracts text from common file types (PDF, plain text, DOCX, XLSX)
and calls robotsix-llmio (OpenRouter transport) to generate summary,
category, and tags via structured output (:class:`EnrichmentModel`).

Images (``image/*``) and scanned/image-based PDFs are handled via a
two-step vision pipeline: a vision-capable model (``enrichment_vision_model``,
by default Gemini 2.0 Flash) first produces a plain-text caption of the image,
then the caption is fed through the text classifier for summary/category/tags.
SVG inputs are rasterized to PNG before being sent to the vision model.
Scanned/image-based PDFs (where pypdf extracts no embedded text) are
rendered to page images via ``pdf2image`` first.

Embeddings go to the dedicated ``embedding.endpoint`` (shared bge-m3
server) — no llmio involvement.

All operations are best-effort — if text extraction fails or the LLM
call errors/times out, enrichment fields are left null rather than
failing the upload.
"""

from __future__ import annotations

import json
import logging
import os
from typing import Any, cast

import httpx
from pydantic import BaseModel, Field
from pydantic_ai.messages import BinaryContent
from robotsix_llmio import get_provider_for_level
from robotsix_llmio.core.factory import get_provider_for_identifier
from robotsix_llmio.core.identifier import parse_model_identifier

from .config import get_settings

# Sentinel returned by :func:`extract_text` for ``image/*`` content types.
# Signals :func:`enrich_file` to use the vision LLM path instead of text.
IMAGE_SENTINEL = "__IMAGE__"

# Sentinel returned by :func:`extract_text` for scanned/image-based PDFs
# where pypdf extracts no embedded text.  Signals :func:`enrich_file` to
# render pages to images and use the vision LLM path.
SCANNED_PDF_SENTINEL = "__SCANNED_PDF__"

logger = logging.getLogger(__name__)

settings = get_settings()


# ── Structured enrichment model ────────────────────────────────────


class EnrichmentModel(BaseModel):
    """Structured output from the LLM enrichment prompt."""

    summary: str = Field(description="A 1-3 sentence summary of the content.")
    category: str = Field(
        description=(
            'A single category label (e.g. "document", "image", "code", '
            '"spreadsheet", "presentation", "legal", "financial", '
            '"scientific", "other").'
        ),
    )
    tags: list[str] = Field(description="Up to 10 keyword tags.", max_length=10)


# ── Text extraction ────────────────────────────────────────────────


def extract_text(content: bytes, content_type: str) -> str | None:
    """Best-effort text extraction from *content* based on its MIME type.

    Returns the extracted text string, or ``None`` if extraction is
    unsupported or fails.
    """
    content_type_lower = content_type.lower()

    # ── Plain text ──────────────────────────────────────────────
    if content_type_lower.startswith("text/"):
        try:
            return content.decode("utf-8", errors="replace")
        except Exception:
            return content.decode("latin-1", errors="replace")

    # ── PDF ─────────────────────────────────────────────────────
    if content_type_lower == "application/pdf":
        try:
            from io import BytesIO

            from pypdf import PdfReader

            reader = PdfReader(BytesIO(content))
            parts: list[str] = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    parts.append(page_text)
            text = "\n".join(parts).strip()
            if text:
                return text
            # Empty text → scanned/image-based PDF
            return SCANNED_PDF_SENTINEL
        except Exception:
            logger.warning("Failed to extract text from PDF", exc_info=True)
            return None

    # ── DOCX (Office Open XML word processing) ──────────────────
    if (
        "officedocument.wordprocessingml" in content_type_lower
        or content_type_lower
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        try:
            from io import BytesIO

            from docx import Document

            doc = Document(BytesIO(content))
            text = "\n".join(p.text for p in doc.paragraphs).strip()
            return text or None
        except Exception:
            logger.warning("Failed to extract text from DOCX", exc_info=True)
            return None

    # ── XLSX / spreadsheet ──────────────────────────────────────
    if "spreadsheet" in content_type_lower or "excel" in content_type_lower:
        try:
            from io import BytesIO

            from openpyxl import load_workbook  # type: ignore[import-untyped]

            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            text = "\n".join(rows).strip()
            wb.close()
            return text or None
        except Exception:
            logger.warning("Failed to extract text from XLSX", exc_info=True)
            return None

    # ── Images ──────────────────────────────────────────────────────
    if content_type_lower.startswith("image/"):
        return IMAGE_SENTINEL

    return None


# ── SVG rasterization helper ──────────────────────────────────────


def _rasterize_svg(content: bytes) -> bytes:
    """Rasterize an SVG document to PNG bytes so the vision model can read it.

    Vision models receive raster ``image/*`` bytes; an ``image/svg+xml``
    upload must first be converted to a bitmap.  Uses ``cairosvg`` (requires
    the ``libcairo2`` system library).  Raises ``Exception`` on failure
    (best-effort — caller catches).
    """
    import cairosvg

    return cairosvg.svg2png(bytestring=content)


# ── Scanned PDF helpers ────────────────────────────────────────────


def _render_pdf_pages(content: bytes) -> list[bytes]:
    """Render each page of a PDF to PNG bytes using pdf2image.

    Requires the ``poppler`` system library to be installed.
    Raises ``Exception`` on failure (best-effort — caller catches).
    """
    from io import BytesIO

    from pdf2image import convert_from_bytes

    images = convert_from_bytes(content, fmt="png")
    page_images: list[bytes] = []
    for img in images:
        buf = BytesIO()
        img.save(buf, format="PNG")
        page_images.append(buf.getvalue())
    return page_images


def _merge_page_results(page_results: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge enrichment results from multiple PDF pages.

    Combines summaries by joining, takes the first category, and
    deduplicates tags (preserving order, capped at 10).
    """
    if not page_results:
        return {"summary": "", "category": None, "tags": []}

    summaries = [r.get("summary", "") for r in page_results if r.get("summary")]
    categories = [r.get("category") for r in page_results if r.get("category")]
    all_tags: list[str] = []
    for r in page_results:
        all_tags.extend(r.get("tags", []))

    seen: set[str] = set()
    unique_tags: list[str] = []
    for tag in all_tags:
        if tag not in seen:
            seen.add(tag)
            unique_tags.append(tag)

    return {
        "summary": " ".join(summaries),
        "category": categories[0] if categories else None,
        "tags": unique_tags[:10],
    }


# ── Langfuse environment wiring ────────────────────────────────────


def _wire_langfuse_env() -> None:
    """Export Langfuse credentials into the process environment.

    llmio's Langfuse export activates when ``LANGFUSE_PUBLIC_KEY``,
    ``LANGFUSE_SECRET_KEY``, and ``LANGFUSE_BASE_URL`` are set.  We
    read them from the canonical ``langfuse`` config block for the
    ``robotsix-file-hub`` project alias.
    """
    lf = settings.langfuse
    project = lf.projects.get("robotsix-file-hub")
    if project is None:
        return

    secret = project.secret_key.get_secret_value()
    if not project.public_key or not secret:
        return

    os.environ.setdefault("LANGFUSE_PUBLIC_KEY", project.public_key)
    os.environ.setdefault("LANGFUSE_SECRET_KEY", secret)
    os.environ.setdefault("LANGFUSE_BASE_URL", lf.host)


# ── LLM call (chat enrichment via robotsix-llmio) ───────────────────


async def call_llm(text: str) -> dict[str, Any]:
    """Call robotsix-llmio for structured enrichment.

    Builds a provider and agent at the configured ``enrichment_llm_tier_level``
    (default 1) with ``output_type=EnrichmentModel``.  PromptedOutput wraps the
    model for JSON-in-text structured output — no hand-rolled fence parsing.

    Raises ``Exception`` on failure (best-effort — caller catches).
    """
    _wire_langfuse_env()

    level = settings.enrichment_llm_tier_level

    # Resolve the OpenRouter API key for the robotsix-file-hub alias
    or_key = settings.openrouter.keys.get("robotsix-file-hub")
    api_key: str | None = None
    if or_key is not None:
        api_key = or_key.get_secret_value()

    provider = get_provider_for_level(level, api_key=api_key)

    prompt = (
        "Analyze the following text content from a file and return a JSON object "
        "with exactly three fields:\n"
        '  "summary": a 1-3 sentence summary of the content,\n'
        '  "category": a single category label (e.g. "document", "image", "code", '
        '"spreadsheet", "presentation", "legal", "financial", "scientific", "other"),\n'
        '  "tags": a list of up to 10 keyword tags (strings).\n'
        "Respond with only the JSON object, no other text.\n\n"
        f"TEXT:\n{text[:8000]}"
    )

    agent = provider.build_agent(
        level=level,
        system_prompt=prompt,
        output_type=EnrichmentModel,
        name="file-hub-enricher",
        retries=2,
        builtin_tools=False,
        web_tools=False,
    )

    async def _run() -> EnrichmentModel:
        result = await agent.run("")
        return cast(EnrichmentModel, result.output)

    model_result = cast(
        EnrichmentModel,
        await provider.call_with_retry(
            _run,
            what="LLM chat enrichment",
        ),
    )

    return {
        "summary": model_result.summary or "",
        "category": model_result.category or "",
        "tags": model_result.tags or [],
    }


async def _vision_caption(image_bytes: bytes, content_type: str) -> str:
    """Call the configured vision model to produce a plain-text caption.

    Sends *image_bytes* as :class:`BinaryContent` multimodal input to the
    model bound by the ``enrichment_vision_model`` setting and returns the
    model's raw caption (``str``) — no structured fields.

    Raises ``Exception`` on failure (best-effort — caller catches).
    """
    _wire_langfuse_env()

    or_key = settings.openrouter.keys.get("robotsix-file-hub")
    api_key: str | None = None
    if or_key is not None:
        api_key = or_key.get_secret_value()

    identifier = settings.enrichment_vision_model
    provider = get_provider_for_identifier(identifier, api_key=api_key)
    model_name = parse_model_identifier(identifier).model_name

    prompt = (
        "Analyze the image and write a concise, factual caption describing "
        "what is visible. Respond with only the caption — a single paragraph "
        "of 1-3 sentences — with no JSON, labels, or extra commentary."
    )

    agent = provider.build_agent(
        level=1,
        model=model_name,
        system_prompt=prompt,
        output_type=str,
        name="file-hub-vision-captioner",
        retries=2,
        builtin_tools=False,
        web_tools=False,
    )

    image_content = BinaryContent(data=image_bytes, media_type=content_type)

    async def _run() -> str:
        result = await agent.run([image_content])
        return cast(str, result.output)

    return cast(
        str,
        await provider.call_with_retry(
            _run,
            what="LLM vision caption",
        ),
    )


async def call_llm_vision(image_bytes: bytes, content_type: str) -> dict[str, Any]:
    """Enrich a raster image (or scanned-PDF page) via a two-step pipeline.

    Step 1 drives the image through the configured ``enrichment_vision_model``
    to obtain a plain-text caption.  SVG inputs are rasterized to PNG first
    (vision models read bitmaps).  Step 2 feeds that caption to the text
    classifier (:func:`call_llm`) which produces the standard
    ``EnrichmentModel`` summary/category/tags.

    Returns the same ``EnrichmentModel`` fields as :func:`call_llm`.

    Raises ``Exception`` on failure (best-effort — caller catches).
    """
    content_type_lower = content_type.lower()
    if content_type_lower in {"image/svg", "image/svg+xml", "application/svg+xml"}:
        image_bytes = _rasterize_svg(image_bytes)
        content_type = "image/png"

    caption = await _vision_caption(image_bytes, content_type)
    return await call_llm(caption)


async def generate_embedding(text: str) -> list[float] | None:
    """Call the dedicated OpenAI-compatible embeddings endpoint.

    Uses the ``embedding`` config block (``embedding.endpoint`` + ``embedding.model``).
    Emits 1024-dim vectors — no llmio involvement.

    Returns ``None`` if the API call fails (best-effort).
    """
    emb = settings.embedding

    headers: dict[str, str] = {"Content-Type": "application/json"}
    api_key = emb.api_key.get_secret_value()
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    try:
        async with httpx.AsyncClient(timeout=emb.timeout) as client:
            response = await client.post(
                f"{emb.endpoint}/embeddings",
                headers=headers,
                json={"model": emb.model, "input": text[:8000]},
            )
            response.raise_for_status()
            data = response.json()
        return list(data["data"][0]["embedding"])
    except Exception:
        logger.warning("Embedding generation failed", exc_info=True)
        return None


# ── Orchestration ───────────────────────────────────────────────────


async def enrich_file(content: bytes, content_type: str) -> dict[str, str | None]:
    """Extract text and call the LLM for enrichment.

    Returns a dict with ``category``, ``tags`` (comma-separated),
    ``summary``, and ``embedding`` (JSON-serialised list of floats)
    — all nullable.  If text extraction yields nothing or the LLM
    call fails, fields are returned as ``None`` (best-effort).
    """
    text = extract_text(content, content_type)
    if not text:
        logger.info(
            "No text extracted for content_type=%s, skipping LLM enrichment",
            content_type,
        )
        return {"category": None, "tags": None, "summary": None, "embedding": None}

    try:
        if text == IMAGE_SENTINEL:
            llm_result = await call_llm_vision(content, content_type)
        elif text == SCANNED_PDF_SENTINEL:
            page_images = _render_pdf_pages(content)
            page_results = []
            for page_bytes in page_images:
                result = await call_llm_vision(page_bytes, "image/png")
                page_results.append(result)
            llm_result = _merge_page_results(page_results)
        else:
            llm_result = await call_llm(text)
        summary_text = llm_result.get("summary", "")
        category = llm_result["category"] or None
        tags = ",".join(llm_result["tags"]) or None
    except Exception:
        logger.warning("LLM enrichment failed for content_type=%s", content_type, exc_info=True)
        category = None
        tags = None
        summary_text = ""

    # Generate embedding from a composite of the enrichment fields
    embedding: str | None = None
    embedding_input = _embedding_input_text(summary_text, category, tags)
    if embedding_input.strip():
        vec = await generate_embedding(embedding_input)
        if vec is not None:
            embedding = json.dumps(vec)

    return {
        "category": category,
        "tags": tags,
        "summary": summary_text or None,
        "embedding": embedding,
    }


def _embedding_input_text(summary: str, category: str | None, tags: str | None) -> str:
    """Build a text string from enrichment fields for embedding generation."""
    parts: list[str] = []
    if summary:
        parts.append(summary)
    if category:
        parts.append(category)
    if tags:
        parts.append(tags.replace(",", " "))
    return " ".join(parts)
